"""
Модуль «Зарплата и кадры» для БухАгент КР.

Налоговые ставки (Кыргызстан):
  Резиденты:    ПН 10% | ПФР 8% (сч. 3531) | ГНПФР 2% (сч. 3534) | СФ работодателя 17.5%
  Нерезиденты:  ПН 10% | ПФР = 0% | ГНПФР = 0%

Проводки при начислении зарплаты:
  Дт 8010 / Кт 3520 — начисление зарплаты (gross)
  Дт 3520 / Кт 3410 — удержан подоходный налог (ПН)
  Дт 3520 / Кт 3531 — удержан ПФР 8% (пенсионный фонд)
  Дт 3520 / Кт 3534 — удержан ГНПФР 2% (накопительный пенсионный фонд)
  Дт 8020 / Кт 3530 — соцфонд работодателя 17.5%
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, date
import io
from database import get_db
from routers.auth import get_current_user
import models
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter()

# ── Ставки налогов КР ──────────────────────────────────────────────────────
TAX = {
    "resident": {
        "income_tax":  0.10,
        "pfr":         0.08,   # ПФР — Пенсионный фонд         (сч. 3531)
        "gnpfr":       0.02,   # ГНПФР — накопит. пенсионный   (сч. 3534)
        "sf_employer": 0.175,
    },
    "foreign": {
        "income_tax":  0.10,
        "pfr":         0.0,    # нерезиденты освобождены от обязательного СФ
        "gnpfr":       0.0,
        "sf_employer": 0.0,
    },
}

MONTH_RU = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]


# ── Pydantic схемы ─────────────────────────────────────────────────────────
class EmployeeCreate(BaseModel):
    full_name:  str
    inn:        Optional[str] = None
    position:   Optional[str] = None
    department: Optional[str] = None
    salary:     float
    hire_date:  str            # YYYY-MM-DD
    is_foreign: bool = False

class EmployeeUpdate(BaseModel):
    full_name:  Optional[str]   = None
    inn:        Optional[str]   = None
    position:   Optional[str]   = None
    department: Optional[str]   = None
    salary:     Optional[float] = None
    is_foreign: Optional[bool]  = None

class AdjustmentItem(BaseModel):
    employee_id: int
    bonus:       float = 0.0   # премия — облагается налогом
    deduction:   float = 0.0   # удержание — не влияет на налог (аванс, штраф и т.д.)

class LeaveCreate(BaseModel):
    employee_id: int
    leave_type:  str            # vacation | sick
    start_date:  str            # YYYY-MM-DD
    end_date:    str            # YYYY-MM-DD
    notes:       Optional[str] = None

class RunPayrollRequest(BaseModel):
    year:        int
    month:       int   # 1-12
    adjustments: list[AdjustmentItem] = []

class PayRequest(BaseModel):
    pay_date:     str = ""      # YYYY-MM-DD, пусто = сегодня
    account_type: str = "bank"  # bank | cash

class AdvanceRequest(BaseModel):
    amount:       float
    pay_date:     str = ""
    account_type: str = "bank"


# ── Хелперы ────────────────────────────────────────────────────────────────
def _calc(emp: models.Employee, bonus: float = 0.0, deduction: float = 0.0) -> dict:
    rates      = TAX["foreign"] if emp.is_foreign else TAX["resident"]
    gross      = emp.salary
    taxable    = round(gross + bonus, 2)         # премия облагается налогом
    income_tax = round(taxable * rates["income_tax"], 2)
    pfr        = round(taxable * rates["pfr"],        2)   # ПФР 8%
    gnpfr      = round(taxable * rates["gnpfr"],      2)   # ГНПФР 2%
    sf_er      = round(taxable * rates["sf_employer"], 2)
    net        = round(taxable - income_tax - pfr - gnpfr - deduction, 2)
    return {
        "employee_id":    emp.id,
        "employee_name":  emp.full_name,
        "position":       emp.position or "",
        "department":     emp.department or "",
        "is_foreign":     emp.is_foreign,
        "gross":          gross,
        "bonus":          round(bonus, 2),
        "deduction":      round(deduction, 2),
        "taxable":        taxable,
        "income_tax":     income_tax,
        "sf_employee":    pfr,     # ПФР — хранится в поле sf_employee
        "gnpfr_employee": gnpfr,   # ГНПФР — новое поле
        "sf_employer":    sf_er,
        "net":            net,
    }


def _emp_dict(emp: models.Employee) -> dict:
    return {
        "id":         emp.id,
        "full_name":  emp.full_name,
        "inn":        emp.inn,
        "position":   emp.position,
        "department": emp.department,
        "salary":     emp.salary,
        "hire_date":  emp.hire_date.strftime("%Y-%m-%d") if emp.hire_date else None,
        "fire_date":  emp.fire_date.strftime("%Y-%m-%d") if emp.fire_date else None,
        "is_foreign": emp.is_foreign,
        "is_active":  emp.is_active,
    }


def _run_dict(run: models.PayrollRun) -> dict:
    return {
        "id":               run.id,
        "year":             run.year,
        "month":            run.month,
        "month_name":       MONTH_RU[run.month] if 1 <= run.month <= 12 else "",
        "status":           run.status,
        "gross_total":       run.gross_total,
        "income_tax_total":  run.income_tax_total,
        "sf_employee_total": run.sf_employee_total,  # ПФР 8%
        "gnpfr_total":       run.gnpfr_total or 0,   # ГНПФР 2%
        "sf_employer_total": run.sf_employer_total,
        "net_total":         run.net_total,
        "is_paid":           run.is_paid or False,
        "paid_at":           run.paid_at.isoformat() if run.paid_at else None,
        "is_tax_paid":       run.is_tax_paid or False,
        "tax_paid_at":       run.tax_paid_at.isoformat() if run.tax_paid_at else None,
        "advance_total":     run.advance_total or 0,
        "is_advance_paid":   run.is_advance_paid or False,
        "advance_paid_at":   run.advance_paid_at.isoformat() if run.advance_paid_at else None,
        "created_at":        run.created_at.isoformat() if run.created_at else None,
    }


def _run_detail(run: models.PayrollRun) -> dict:
    d = _run_dict(run)
    d["entries"] = [
        {
            "id":            e.id,
            "employee_id":   e.employee_id,
            "employee_name": e.employee_name,
            "position":      e.position,
            "department":    e.department or "",
            "is_foreign":    e.is_foreign,
            "gross":         e.gross,
            "bonus":         e.bonus or 0,
            "deduction":     e.deduction or 0,
            "taxable":       e.taxable or e.gross,
            "income_tax":     e.income_tax,
            "sf_employee":    e.sf_employee,   # ПФР 8%
            "gnpfr_employee": e.gnpfr_employee or 0,  # ГНПФР 2%
            "sf_employer":    e.sf_employer,
            "net":            e.net,
        }
        for e in run.entries
    ]
    return d


def _acc_name(code: str, db: Session) -> str:
    acc = db.query(models.ChartOfAccount).filter(
        models.ChartOfAccount.code == code
    ).first()
    return acc.name if acc else code


def _post_entry(company_id: int, debit: str, credit: str, amount: float,
                desc: str, run: models.PayrollRun, db: Session) -> None:
    if amount <= 0:
        return
    entry_date = date(run.year, run.month, 1)
    e = models.JournalEntry(
        company_id         = company_id,
        document_id        = None,
        entry_date         = entry_date,
        debit_account      = debit,
        credit_account     = credit,
        debit_account_name = _acc_name(debit,  db),
        credit_account_name= _acc_name(credit, db),
        amount             = amount,
        currency           = "KGS",
        description        = desc,
        status             = "posted",
        ai_confidence      = 100,
        ai_reasoning       = "Авто-проводка: расчёт зарплаты",
    )
    db.add(e)


# ── Сотрудники ─────────────────────────────────────────────────────────────
@router.get("/{company_id}/employees")
def list_employees(company_id: int,
                   db: Session = Depends(get_db),
                   user = Depends(get_current_user)):
    emps = db.query(models.Employee).filter(
        models.Employee.company_id == company_id
    ).order_by(models.Employee.is_active.desc(),
               models.Employee.full_name).all()
    return [_emp_dict(e) for e in emps]


@router.post("/{company_id}/employees")
def add_employee(company_id: int, data: EmployeeCreate,
                 db: Session = Depends(get_db),
                 user = Depends(get_current_user)):
    hire_date = datetime.strptime(data.hire_date, "%Y-%m-%d")
    emp = models.Employee(
        company_id = company_id,
        full_name  = data.full_name,
        inn        = data.inn,
        position   = data.position,
        department = data.department,
        salary     = data.salary,
        hire_date  = hire_date,
        is_foreign = data.is_foreign,
        is_active  = True,
    )
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return _emp_dict(emp)


@router.patch("/{company_id}/employees/{emp_id}/fire")
def fire_employee(company_id: int, emp_id: int,
                  db: Session = Depends(get_db),
                  user = Depends(get_current_user)):
    emp = db.query(models.Employee).filter(
        models.Employee.id == emp_id,
        models.Employee.company_id == company_id,
    ).first()
    if not emp:
        raise HTTPException(404, "Сотрудник не найден")
    emp.is_active  = False
    emp.fire_date  = datetime.utcnow()
    db.commit()
    return _emp_dict(emp)


@router.delete("/{company_id}/employees/{emp_id}")
def delete_employee(company_id: int, emp_id: int,
                    db: Session = Depends(get_db),
                    user = Depends(get_current_user)):
    emp = db.query(models.Employee).filter(
        models.Employee.id == emp_id,
        models.Employee.company_id == company_id,
    ).first()
    if not emp:
        raise HTTPException(404, "Сотрудник не найден")
    db.delete(emp)
    db.commit()
    return {"ok": True}


# ── Предпросмотр без сохранения ────────────────────────────────────────────
@router.get("/{company_id}/payroll")
def preview_payroll(company_id: int,
                    db: Session = Depends(get_db),
                    user = Depends(get_current_user)):
    """Предварительный расчёт — без сохранения."""
    emps = db.query(models.Employee).filter(
        models.Employee.company_id == company_id,
        models.Employee.is_active  == True,
    ).all()
    rows = [_calc(e) for e in emps]
    return {
        "rows": rows,
        "totals": {
            "gross":          round(sum(r["gross"]          for r in rows), 2),
            "bonus":          round(sum(r["bonus"]          for r in rows), 2),
            "deduction":      round(sum(r["deduction"]      for r in rows), 2),
            "taxable":        round(sum(r["taxable"]        for r in rows), 2),
            "income_tax":     round(sum(r["income_tax"]     for r in rows), 2),
            "sf_employee":    round(sum(r["sf_employee"]    for r in rows), 2),  # ПФР
            "gnpfr_employee": round(sum(r["gnpfr_employee"] for r in rows), 2),  # ГНПФР
            "sf_employer":    round(sum(r["sf_employer"]    for r in rows), 2),
            "net":            round(sum(r["net"]            for r in rows), 2),
        },
        "tax_info": {
            "resident":   TAX["resident"],
            "foreign":    TAX["foreign"],
        }
    }


# ── История расчётов ───────────────────────────────────────────────────────
@router.get("/{company_id}/payroll/history")
def payroll_history(company_id: int,
                    db: Session = Depends(get_db),
                    user = Depends(get_current_user)):
    runs = db.query(models.PayrollRun).filter(
        models.PayrollRun.company_id == company_id
    ).order_by(models.PayrollRun.year.desc(),
               models.PayrollRun.month.desc()).all()
    return [_run_dict(r) for r in runs]


@router.get("/{company_id}/payroll/run/{run_id}")
def get_run(company_id: int, run_id: int,
            db: Session = Depends(get_db),
            user = Depends(get_current_user)):
    run = db.query(models.PayrollRun).filter(
        models.PayrollRun.id         == run_id,
        models.PayrollRun.company_id == company_id,
    ).first()
    if not run:
        raise HTTPException(404, "Расчёт не найден")
    return _run_detail(run)


# ── Провести расчёт ────────────────────────────────────────────────────────
@router.post("/{company_id}/payroll/run")
def run_payroll(company_id: int, data: RunPayrollRequest,
                db: Session = Depends(get_db),
                user = Depends(get_current_user)):
    """Рассчитать, сохранить и создать проводки в журнале."""
    if not (1 <= data.month <= 12):
        raise HTTPException(400, "Неверный месяц")

    # Проверка дубля
    dup = db.query(models.PayrollRun).filter(
        models.PayrollRun.company_id == company_id,
        models.PayrollRun.year       == data.year,
        models.PayrollRun.month      == data.month,
    ).first()
    if dup:
        raise HTTPException(
            400,
            f"Расчёт за {MONTH_RU[data.month]} {data.year} уже проведён. "
            f"Удалите предыдущий, чтобы пересчитать."
        )

    emps = db.query(models.Employee).filter(
        models.Employee.company_id == company_id,
        models.Employee.is_active  == True,
    ).all()
    if not emps:
        raise HTTPException(400, "Нет активных сотрудников")

    # Собираем map adjustments по employee_id
    adj_map = {a.employee_id: a for a in (data.adjustments or [])}

    rows = [
        _calc(e,
              bonus     = adj_map[e.id].bonus     if e.id in adj_map else 0.0,
              deduction = adj_map[e.id].deduction if e.id in adj_map else 0.0)
        for e in emps
    ]

    gross      = round(sum(r["gross"]          for r in rows), 2)
    taxable_t  = round(sum(r["taxable"]        for r in rows), 2)
    it_total   = round(sum(r["income_tax"]     for r in rows), 2)
    pfr_t      = round(sum(r["sf_employee"]    for r in rows), 2)   # ПФР 8%
    gnpfr_t    = round(sum(r["gnpfr_employee"] for r in rows), 2)   # ГНПФР 2%
    sf_er_t    = round(sum(r["sf_employer"]    for r in rows), 2)
    net        = round(sum(r["net"]            for r in rows), 2)

    # Сохраняем PayrollRun
    run = models.PayrollRun(
        company_id        = company_id,
        year              = data.year,
        month             = data.month,
        status            = "posted",
        gross_total       = gross,
        income_tax_total  = it_total,
        sf_employee_total = pfr_t,     # ПФР 8%
        gnpfr_total       = gnpfr_t,   # ГНПФР 2%
        sf_employer_total = sf_er_t,
        net_total         = net,
    )
    db.add(run)
    db.flush()   # нужен run.id

    # Строки по сотрудникам
    for r in rows:
        db.add(models.PayrollRunEntry(
            run_id         = run.id,
            employee_id    = r["employee_id"],
            employee_name  = r["employee_name"],
            position       = r["position"],
            department     = r["department"],
            is_foreign     = r["is_foreign"],
            bonus          = r["bonus"],
            deduction      = r["deduction"],
            gross          = r["gross"],
            taxable        = r["taxable"],
            income_tax     = r["income_tax"],
            sf_employee    = r["sf_employee"],    # ПФР 8%
            gnpfr_employee = r["gnpfr_employee"], # ГНПФР 2%
            sf_employer    = r["sf_employer"],
            net            = r["net"],
        ))

    # Проводки
    label = f"{MONTH_RU[data.month]} {data.year}"
    _post_entry(company_id, "8010", "3520", gross,
                f"Начисление заработной платы за {label}", run, db)
    _post_entry(company_id, "3520", "3410", it_total,
                f"Удержан подоходный налог (ПН 10%) за {label}", run, db)
    _post_entry(company_id, "3520", "3531", pfr_t,
                f"Удержан ПФР (8%) за {label}", run, db)
    _post_entry(company_id, "3520", "3534", gnpfr_t,
                f"Удержан ГНПФР (2%) за {label}", run, db)
    _post_entry(company_id, "8020", "3530", sf_er_t,
                f"Начислен соцфонд работодателя (17.5%) за {label}", run, db)

    db.commit()
    db.refresh(run)
    return _run_detail(run)


# ── Удалить расчёт ─────────────────────────────────────────────────────────
@router.delete("/{company_id}/payroll/run/{run_id}")
def delete_run(company_id: int, run_id: int,
               db: Session = Depends(get_db),
               user = Depends(get_current_user)):
    run = db.query(models.PayrollRun).filter(
        models.PayrollRun.id         == run_id,
        models.PayrollRun.company_id == company_id,
    ).first()
    if not run:
        raise HTTPException(404, "Расчёт не найден")
    db.delete(run)
    db.commit()
    return {"ok": True}


# ── Аванс ─────────────────────────────────────────────────────────────────
@router.post("/{company_id}/payroll/run/{run_id}/advance")
def pay_advance(company_id: int, run_id: int, data: AdvanceRequest,
                db: Session = Depends(get_db),
                user = Depends(get_current_user)):
    """
    Зафиксировать ранее выплаченный аванс — Дт 3520 / Кт 1210 (1110).
    Уменьшает итоговую сумму при выплате основной зарплаты.
    """
    run = db.query(models.PayrollRun).filter(
        models.PayrollRun.id         == run_id,
        models.PayrollRun.company_id == company_id,
    ).first()
    if not run:
        raise HTTPException(404, "Расчёт не найден")
    if run.is_advance_paid:
        raise HTTPException(400, "Аванс уже зафиксирован")

    credit  = "1110" if data.account_type == "cash" else "1210"
    pay_dt  = datetime.strptime(data.pay_date, "%Y-%m-%d") if data.pay_date else datetime.utcnow()
    label   = f"{MONTH_RU[run.month]} {run.year}"

    _post_entry(company_id, "3520", credit, data.amount,
                f"Аванс по заработной плате за {label}", run, db)

    run.advance_total   = round(data.amount, 2)
    run.is_advance_paid = True
    run.advance_paid_at = pay_dt
    db.commit()
    db.refresh(run)
    return _run_detail(run)


# ── Редактировать сотрудника ────────────────────────────────────────────────
@router.patch("/{company_id}/employees/{emp_id}")
def update_employee(company_id: int, emp_id: int, data: EmployeeUpdate,
                    db: Session = Depends(get_db),
                    user = Depends(get_current_user)):
    emp = db.query(models.Employee).filter(
        models.Employee.id         == emp_id,
        models.Employee.company_id == company_id,
    ).first()
    if not emp:
        raise HTTPException(404, "Сотрудник не найден")
    if data.full_name  is not None: emp.full_name  = data.full_name
    if data.inn        is not None: emp.inn        = data.inn
    if data.position   is not None: emp.position   = data.position
    if data.department is not None: emp.department = data.department
    if data.salary     is not None: emp.salary     = data.salary
    if data.is_foreign is not None: emp.is_foreign = data.is_foreign
    db.commit()
    return _emp_dict(emp)


# ── Выплатить зарплату ─────────────────────────────────────────────────────
@router.post("/{company_id}/payroll/run/{run_id}/pay")
def pay_salary(company_id: int, run_id: int, data: PayRequest,
               db: Session = Depends(get_db),
               user = Depends(get_current_user)):
    """
    Выплатить зарплату — создаёт проводку Дт 3520 / Кт 1210 (банк) или Кт 1110 (касса).
    """
    run = db.query(models.PayrollRun).filter(
        models.PayrollRun.id         == run_id,
        models.PayrollRun.company_id == company_id,
    ).first()
    if not run:
        raise HTTPException(404, "Расчёт не найден")
    if run.is_paid:
        raise HTTPException(400, "Зарплата уже выплачена")

    credit    = "1110" if data.account_type == "cash" else "1210"
    pay_dt    = datetime.strptime(data.pay_date, "%Y-%m-%d") if data.pay_date else datetime.utcnow()
    label     = f"{MONTH_RU[run.month]} {run.year}"
    advance   = run.advance_total or 0
    remaining = round(run.net_total - advance, 2)

    if remaining > 0:
        _post_entry(company_id, "3520", credit, remaining,
                    f"Выплата заработной платы за {label}"
                    + (f" (за вычетом аванса {advance:,.2f})" if advance else ""),
                    run, db)

    run.is_paid = True
    run.paid_at = pay_dt
    db.commit()
    db.refresh(run)
    return _run_detail(run)


# ── Оплатить налоги в бюджет ───────────────────────────────────────────────
@router.post("/{company_id}/payroll/run/{run_id}/pay-taxes")
def pay_taxes(company_id: int, run_id: int, data: PayRequest,
              db: Session = Depends(get_db),
              user = Depends(get_current_user)):
    """
    Оплата налогов и соцфонда в бюджет:
      Дт 3410 / Кт 1210 — подоходный налог
      Дт 3530 / Кт 1210 — социальный фонд (работник + работодатель)
    """
    run = db.query(models.PayrollRun).filter(
        models.PayrollRun.id         == run_id,
        models.PayrollRun.company_id == company_id,
    ).first()
    if not run:
        raise HTTPException(404, "Расчёт не найден")
    if run.is_tax_paid:
        raise HTTPException(400, "Налоги уже оплачены")

    credit    = "1110" if data.account_type == "cash" else "1210"
    pay_dt    = datetime.strptime(data.pay_date, "%Y-%m-%d") if data.pay_date else datetime.utcnow()
    label     = f"{MONTH_RU[run.month]} {run.year}"
    pfr_amt   = run.sf_employee_total or 0
    gnpfr_amt = run.gnpfr_total or 0
    sf_er_amt = run.sf_employer_total or 0

    _post_entry(company_id, "3410", credit, run.income_tax_total,
                f"Оплата подоходного налога (ПН) за {label}", run, db)
    _post_entry(company_id, "3531", credit, pfr_amt,
                f"Оплата ПФР (8%) за {label}", run, db)
    _post_entry(company_id, "3534", credit, gnpfr_amt,
                f"Оплата ГНПФР (2%) за {label}", run, db)
    _post_entry(company_id, "3530", credit, sf_er_amt,
                f"Оплата соцфонда работодателя за {label}", run, db)

    run.is_tax_paid = True
    run.tax_paid_at = pay_dt
    db.commit()
    db.refresh(run)
    return _run_detail(run)


# ── Отпуска и больничные ───────────────────────────────────────────────────
def _leave_dict(leave: models.EmployeeLeave, emp_name: str = "") -> dict:
    return {
        "id":            leave.id,
        "employee_id":   leave.employee_id,
        "employee_name": emp_name,
        "leave_type":    leave.leave_type,
        "start_date":    leave.start_date.isoformat() if leave.start_date else None,
        "end_date":      leave.end_date.isoformat()   if leave.end_date   else None,
        "days":          leave.days,
        "daily_rate":    leave.daily_rate,
        "pay_amount":    leave.pay_amount,
        "notes":         leave.notes,
        "created_at":    leave.created_at.isoformat() if leave.created_at else None,
    }


@router.get("/{company_id}/leaves")
def list_leaves(company_id: int,
                db:   Session = Depends(get_db),
                user  = Depends(get_current_user)):
    leaves = db.query(models.EmployeeLeave).filter(
        models.EmployeeLeave.company_id == company_id
    ).order_by(models.EmployeeLeave.start_date.desc()).all()

    emp_ids = list({l.employee_id for l in leaves})
    emp_map: dict = {}
    if emp_ids:
        emps    = db.query(models.Employee).filter(models.Employee.id.in_(emp_ids)).all()
        emp_map = {e.id: e.full_name for e in emps}

    return [_leave_dict(l, emp_map.get(l.employee_id, "")) for l in leaves]


@router.post("/{company_id}/leaves")
def create_leave(company_id: int, data: LeaveCreate,
                 db:   Session = Depends(get_db),
                 user  = Depends(get_current_user)):
    emp = db.query(models.Employee).filter(
        models.Employee.id         == data.employee_id,
        models.Employee.company_id == company_id,
    ).first()
    if not emp:
        raise HTTPException(404, "Сотрудник не найден")

    start = date.fromisoformat(data.start_date)
    end   = date.fromisoformat(data.end_date)
    if end < start:
        raise HTTPException(400, "Дата окончания раньше даты начала")

    days       = (end - start).days + 1
    daily_rate = round(emp.salary / 25, 2)

    if data.leave_type == "vacation":
        pay_amount = round(daily_rate * days, 2)
        desc = f"Отпускные: {emp.full_name}, {days} дн. ({start} – {end})"
    elif data.leave_type == "sick":
        employer_days = min(3, days)
        foms_days     = max(0, days - 3)
        pay_amount    = round(daily_rate * employer_days, 2)
        foms_note     = f", дни 4–{days} ({foms_days} дн.) — ФОМС" if foms_days else ""
        desc = (f"Больничный: {emp.full_name}, {days} дн. "
                f"(работодатель за {employer_days} дн.{foms_note})")
    else:
        pay_amount = round(daily_rate * days, 2)
        desc = f"Начисление: {emp.full_name}, {days} дн."

    leave = models.EmployeeLeave(
        company_id  = company_id,
        employee_id = emp.id,
        leave_type  = data.leave_type,
        start_date  = start,
        end_date    = end,
        days        = days,
        daily_rate  = daily_rate,
        pay_amount  = pay_amount,
        notes       = data.notes,
    )
    db.add(leave)
    db.flush()

    if pay_amount > 0:
        je = models.JournalEntry(
            company_id          = company_id,
            document_id         = None,
            entry_date          = start,
            debit_account       = "8010",
            credit_account      = "3520",
            debit_account_name  = _acc_name("8010", db),
            credit_account_name = _acc_name("3520", db),
            amount              = pay_amount,
            currency            = "KGS",
            description         = desc,
            status              = "posted",
            ai_confidence       = 100,
            ai_reasoning        = "Авто-проводка: отпуск/больничный",
        )
        db.add(je)
        db.flush()
        leave.journal_entry_id = je.id

    db.commit()
    db.refresh(leave)
    return _leave_dict(leave, emp.full_name)


@router.delete("/{company_id}/leaves/{leave_id}")
def delete_leave(company_id: int, leave_id: int,
                 db:   Session = Depends(get_db),
                 user  = Depends(get_current_user)):
    leave = db.query(models.EmployeeLeave).filter(
        models.EmployeeLeave.id         == leave_id,
        models.EmployeeLeave.company_id == company_id,
    ).first()
    if not leave:
        raise HTTPException(404, "Запись не найдена")
    db.delete(leave)
    db.commit()
    return {"ok": True}


# ── Excel-выгрузка расчётной ведомости ─────────────────────────────────────
@router.get("/{company_id}/payroll/run/{run_id}/export")
def export_run_excel(company_id: int, run_id: int,
                     db:   Session = Depends(get_db),
                     user  = Depends(get_current_user)):
    run = db.query(models.PayrollRun).filter(
        models.PayrollRun.id         == run_id,
        models.PayrollRun.company_id == company_id,
    ).first()
    if not run:
        raise HTTPException(404, "Расчёт не найден")

    company = db.query(models.Company).filter(models.Company.id == company_id).first()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH_RU[run.month]} {run.year}"

    # ── Заголовок ──────────────────────────────────────────────────────────
    NUM_COLS = 11
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    ws["A1"] = (f"РАСЧЁТНАЯ ВЕДОМОСТЬ ЗА "
                f"{MONTH_RU[run.month].upper()} {run.year} ГОДА")
    ws["A1"].font      = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    if company:
        ws.merge_cells(f"A2:{get_column_letter(NUM_COLS)}2")
        ws["A2"]           = company.name
        ws["A2"].font      = Font(size=11, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

    # ── Шапка таблицы ──────────────────────────────────────────────────────
    HEADERS = [
        ("№",           5),
        ("ФИО",        30),
        ("Должность",  18),
        ("Оклад",      13),
        ("Премия",     10),
        ("Удержание",  10),
        ("ПН (10%)",   11),
        ("ПФР (8%)",   11),
        ("ГНПФР (2%)", 11),
        ("СФ р-ль",    11),
        ("К выдаче",   13),
    ]
    fill_hdr = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, (hdr, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font      = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.fill      = fill_hdr
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[4].height = 28

    # ── Группировка по подразделениям ─────────────────────────────────────
    NUM_FMT = "#,##0.00"
    gnpfr_total_run = run.gnpfr_total or 0

    # Собираем группы: { dept_name: [entries] }
    from collections import defaultdict
    groups = defaultdict(list)
    no_dept_key = "— Без подразделения —"
    for e in run.entries:
        key = (e.department or "").strip() or no_dept_key
        groups[key].append(e)

    has_departments = any(k != no_dept_key for k in groups)

    fill_dept = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # голубой для подразделения
    fill_tot  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # зелёный для итого

    cur_row = 5
    emp_num = 1

    def _write_num_row(ws, r, num, entry, NUM_FMT, NUM_COLS, indent=False):
        """Записать одну строку сотрудника."""
        gnpfr = entry.gnpfr_employee or 0
        values = [num,
                  ("  " if indent else "") + entry.employee_name,
                  entry.position or "",
                  entry.gross, entry.bonus or 0, entry.deduction or 0,
                  entry.income_tax, entry.sf_employee, gnpfr,
                  entry.sf_employer, entry.net]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            if col >= 4:
                cell.number_format = NUM_FMT
                cell.alignment     = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center" if col == 1 else "left")

    def _sum_entries(entries):
        """Суммирует строки для итоговой строки подразделения."""
        return [
            sum(e.gross         for e in entries),
            sum(e.bonus or 0    for e in entries),
            sum(e.deduction or 0 for e in entries),
            sum(e.income_tax    for e in entries),
            sum(e.sf_employee   for e in entries),
            sum(e.gnpfr_employee or 0 for e in entries),
            sum(e.sf_employer   for e in entries),
            sum(e.net           for e in entries),
        ]

    for dept_name, entries in sorted(groups.items()):
        if has_departments:
            # Шапка подразделения
            ws.merge_cells(f"A{cur_row}:{get_column_letter(NUM_COLS)}{cur_row}")
            ws[f"A{cur_row}"] = dept_name
            ws[f"A{cur_row}"].font      = Font(bold=True, size=11, color="1F3864")
            ws[f"A{cur_row}"].fill      = fill_dept
            ws[f"A{cur_row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

        for e in entries:
            _write_num_row(ws, cur_row, emp_num, e, NUM_FMT, NUM_COLS,
                           indent=has_departments)
            emp_num  += 1
            cur_row  += 1

        # Итого по подразделению (только если их несколько)
        if has_departments and len(entries) > 1:
            dept_sums = _sum_entries(entries)
            ws.merge_cells(f"A{cur_row}:C{cur_row}")
            ws[f"A{cur_row}"]      = f"Итого: {dept_name}"
            ws[f"A{cur_row}"].font = Font(bold=True, italic=True, size=10)
            ws[f"A{cur_row}"].fill = fill_dept
            for col, val in zip(range(4, NUM_COLS + 1), dept_sums):
                cell               = ws.cell(row=cur_row, column=col, value=val)
                cell.font          = Font(bold=True)
                cell.number_format = NUM_FMT
                cell.alignment     = Alignment(horizontal="right")
                cell.fill          = fill_dept
            cur_row += 1

    # ── Итого по всем ──────────────────────────────────────────────────────
    tr = cur_row
    ws.merge_cells(f"A{tr}:C{tr}")
    ws[f"A{tr}"]      = "ИТОГО"
    ws[f"A{tr}"].font = Font(bold=True, size=11)
    totals = [run.gross_total, 0, 0,
              run.income_tax_total, run.sf_employee_total, gnpfr_total_run,
              run.sf_employer_total, run.net_total]
    for col, val in zip(range(4, NUM_COLS + 1), totals):
        cell               = ws.cell(row=tr, column=col, value=val)
        cell.font          = Font(bold=True)
        cell.number_format = NUM_FMT
        cell.alignment     = Alignment(horizontal="right")
        cell.fill          = fill_tot
    ws[f"A{tr}"].fill = fill_tot

    # ── Примечания ─────────────────────────────────────────────────────────
    note_row = tr + 1
    ws.merge_cells(f"A{note_row}:{get_column_letter(NUM_COLS)}{note_row}")
    notes = []
    if run.advance_total:
        notes.append(f"Аванс: {run.advance_total:,.2f} KGS")
    if run.sf_employer_total:
        notes.append(f"СФ работодателя: {run.sf_employer_total:,.2f} KGS (за счёт компании)")
    ws[f"A{note_row}"] = "  •  ".join(notes) if notes else ""
    ws[f"A{note_row}"].font      = Font(italic=True, size=9)
    ws[f"A{note_row}"].alignment = Alignment(horizontal="right")

    # ── Подписи ─────────────────────────────────────────────────────────────
    sig = tr + 3
    ws.cell(row=sig,     column=1,
            value="Руководитель:  ________________________________")
    ws.cell(row=sig + 1, column=1,
            value="Бухгалтер:     ________________________________")

    # ── Отдаём файл ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"payroll_{run.year}_{run.month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
