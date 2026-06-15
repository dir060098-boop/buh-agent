"""
Модуль «ЭСФ» для БухАгент КР.

Входящие (incoming): ЭСФ от поставщиков
  Статусы: pending (не принят) → accepted (принят)

Исходящие (outgoing): ЭСФ выставленные покупателям
  Статусы: pending (не выставлен) → issued (выставлен)
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
import io
from database import get_db
from routers.auth import get_current_user, require_company
import models

router = APIRouter()


# ── Pydantic схемы ─────────────────────────────────────────────────────────
class ESFCreate(BaseModel):
    direction:       str   = "incoming"   # incoming | outgoing
    esf_number:      str
    esf_date:        str                  # YYYY-MM-DD
    supplier_name:   Optional[str] = None
    supplier_inn:    Optional[str] = None
    buyer_name:      Optional[str] = None
    buyer_inn:       Optional[str] = None
    contract_number: Optional[str] = None
    amount:          float
    vat_rate:        str   = "12"         # "12" | "0" | "exempt"
    vat_amount:      float = 0


# ── Хелпер: dict ───────────────────────────────────────────────────────────
def _esf_dict(e: models.ESF) -> dict:
    return {
        "id":                  e.id,
        "company_id":          e.company_id,
        "direction":           e.direction or "incoming",
        "esf_number":          e.esf_number,
        "esf_date":            e.esf_date.isoformat()[:10] if e.esf_date else None,
        "supplier_name":       e.supplier_name,
        "supplier_inn":        e.supplier_inn,
        "buyer_name":          e.buyer_name,
        "buyer_inn":           e.buyer_inn,
        "contract_number":     e.contract_number,
        "amount":              e.amount or 0,
        "vat_amount":          e.vat_amount or 0,
        "vat_rate":            e.vat_rate or "12",
        "status":              e.status or "pending",
        "accepted_at":         e.accepted_at.isoformat() if e.accepted_at else None,
        "linked_document_id":  e.linked_document_id,
        "bank_transaction_id": e.bank_transaction_id,
        "created_at":          e.created_at.isoformat() if e.created_at else None,
    }


# ── Список с фильтрами ─────────────────────────────────────────────────────
@router.get("/{company_id}")
def list_esf(
    company_id: int,
    direction:  Optional[str] = None,
    date_from:  Optional[str] = None,
    date_to:    Optional[str] = None,
    limit:  int = 100,
    offset: int = 0,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    q = db.query(models.ESF).filter(models.ESF.company_id == company_id)
    if direction:
        q = q.filter(models.ESF.direction == direction)
    if date_from:
        q = q.filter(models.ESF.esf_date >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        dt = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        q = q.filter(models.ESF.esf_date <= dt)
    total   = q.count()
    records = q.order_by(models.ESF.esf_date.desc()).offset(offset).limit(limit).all()
    return {"items": [_esf_dict(r) for r in records], "total": total, "has_more": offset + limit < total}


# ── Книга покупок / продаж ─────────────────────────────────────────────────
@router.get("/{company_id}/book")
def get_book(
    company_id: int,
    direction:  str = "incoming",
    date_from:  Optional[str] = None,
    date_to:    Optional[str] = None,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    q = db.query(models.ESF).filter(
        models.ESF.company_id == company_id,
        models.ESF.direction  == direction,
    )
    if date_from:
        q = q.filter(models.ESF.esf_date >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        dt = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        q = q.filter(models.ESF.esf_date <= dt)
    records = q.order_by(models.ESF.esf_date).all()
    items = [_esf_dict(r) for r in records]
    return {
        "direction":       direction,
        "items":           items,
        "total_amount":    round(sum(r.amount     or 0 for r in records), 2),
        "total_vat":       round(sum(r.vat_amount or 0 for r in records), 2),
        "count":           len(records),
        "accepted_count":  sum(1 for r in records if r.status in ("accepted", "issued")),
        "pending_count":   sum(1 for r in records if r.status == "pending"),
    }


# ── Экспорт Книги покупок/продаж в Excel ─────────────────────────────────
@router.get("/{company_id}/book/export")
def export_book(
    company_id: int,
    direction:  str = "incoming",
    date_from:  Optional[str] = None,
    date_to:    Optional[str] = None,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    """Выгружает Книгу покупок или продаж в формате Excel (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import models as _m

    # Получаем компанию для названия
    company = db.query(_m.Company).filter(_m.Company.id == company_id).first()
    company_name = company.name if company else f"Компания #{company_id}"

    q = db.query(_m.ESF).filter(_m.ESF.company_id == company_id, _m.ESF.direction == direction)
    if date_from:
        q = q.filter(_m.ESF.esf_date >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        dt = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        q = q.filter(_m.ESF.esf_date <= dt)
    records = q.order_by(_m.ESF.esf_date).all()

    book_title = "Книга покупок" if direction == "incoming" else "Книга продаж"
    period_str = ""
    if date_from or date_to:
        period_str = f"  {date_from or ''} — {date_to or ''}"

    wb = Workbook()
    ws = wb.active
    ws.title = book_title[:31]

    # Стили
    accent   = "1A56DB"
    hdr_fill = PatternFill("solid", fgColor=accent)
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sub_fill = PatternFill("solid", fgColor="EEF2FF")
    sub_font = Font(name="Arial", bold=True, size=10)
    cell_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", bold=True, size=10)
    total_fill = PatternFill("solid", fgColor="DBEAFE")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовок книги
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{company_name} — {book_title}{period_str}"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, color="888888")
    ws.row_dimensions[2].height = 14

    # Заголовки колонок
    if direction == "incoming":
        headers = ["№", "Дата ЭСФ", "Номер ЭСФ", "Поставщик", "ИНН поставщика",
                   "Номер договора", "Сумма с НДС", "НДС", "Ставка НДС", "Статус"]
    else:
        headers = ["№", "Дата ЭСФ", "Номер ЭСФ", "Покупатель", "ИНН покупателя",
                   "Номер договора", "Сумма с НДС", "НДС", "Ставка НДС", "Статус"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font    = hdr_font
        cell.fill    = hdr_fill
        cell.alignment = hdr_aln
        cell.border  = border
    ws.row_dimensions[3].height = 30

    # Данные
    STATUS_RU = {"pending": "Не принят", "accepted": "Принят", "issued": "Выставлен"}
    VAT_RATE_RU = {"12": "12%", "0": "0%", "exempt": "Без НДС"}

    total_amount = 0.0
    total_vat    = 0.0

    for idx, r in enumerate(records, 1):
        row_n = idx + 3
        date_str = r.esf_date.strftime("%d.%m.%Y") if r.esf_date else "—"
        counterparty = (r.supplier_name if direction == "incoming" else r.buyer_name) or "—"
        inn          = (r.supplier_inn  if direction == "incoming" else r.buyer_inn)  or "—"
        amount  = r.amount     or 0.0
        vat     = r.vat_amount or 0.0
        total_amount += amount
        total_vat    += vat

        row_data = [
            idx, date_str, r.esf_number or "—", counterparty, inn,
            r.contract_number or "—",
            round(amount, 2), round(vat, 2),
            VAT_RATE_RU.get(r.vat_rate or "12", r.vat_rate),
            STATUS_RU.get(r.status, r.status),
        ]
        fill = sub_fill if idx % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            cell.font   = cell_font
            cell.border = border
            if fill:
                cell.fill = fill
            if col in (7, 8):   # суммы
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")

    # Строка итогов
    total_row = len(records) + 4
    ws.cell(total_row, 1, "ИТОГО").font   = total_font
    ws.cell(total_row, 1).fill   = total_fill
    ws.merge_cells(f"A{total_row}:F{total_row}")
    ws.cell(total_row, 1).alignment = Alignment(horizontal="right")

    for col in range(1, 11):
        c = ws.cell(total_row, col)
        c.border = border
        c.fill   = total_fill
        c.font   = total_font

    t_amt = ws.cell(total_row, 7, round(total_amount, 2))
    t_vat = ws.cell(total_row, 8, round(total_vat,    2))
    t_amt.number_format = t_vat.number_format = '#,##0.00'
    t_amt.alignment = t_vat.alignment = Alignment(horizontal="right")

    ws.cell(total_row, 9, f"{len(records)} записей").font = total_font

    # Ширины колонок
    widths = [5, 12, 22, 35, 18, 18, 15, 13, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = book_title.replace(" ", "_")
    filename = f"{safe_title}_{company_id}"
    if date_from: filename += f"_от{date_from}"
    if date_to:   filename += f"_до{date_to}"
    filename += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Создать ────────────────────────────────────────────────────────────────
@router.post("/{company_id}")
def create_esf(
    company_id: int,
    data: ESFCreate,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    esf_date = (datetime.strptime(data.esf_date, "%Y-%m-%d")
                if data.esf_date else datetime.utcnow())

    # Авторасчёт НДС если не указан и ставка 12%
    vat_amount = data.vat_amount
    if vat_amount == 0 and data.vat_rate == "12" and data.amount > 0:
        vat_amount = round(data.amount * 12 / 112, 2)

    esf = models.ESF(
        company_id      = company_id,
        direction       = data.direction,
        esf_number      = data.esf_number,
        esf_date        = esf_date,
        supplier_name   = data.supplier_name,
        supplier_inn    = data.supplier_inn,
        buyer_name      = data.buyer_name,
        buyer_inn       = data.buyer_inn,
        contract_number = data.contract_number,
        amount          = data.amount,
        vat_amount      = vat_amount,
        vat_rate        = data.vat_rate,
        status          = "pending",
    )
    db.add(esf)
    db.commit()
    db.refresh(esf)
    return _esf_dict(esf)


# ── Удалить ────────────────────────────────────────────────────────────────
@router.delete("/{company_id}/{esf_id}")
def delete_esf(
    company_id: int,
    esf_id: int,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    esf = db.query(models.ESF).filter(
        models.ESF.id         == esf_id,
        models.ESF.company_id == company_id,
    ).first()
    if not esf:
        raise HTTPException(404, "ЭСФ не найден")
    db.delete(esf)
    db.commit()
    return {"ok": True}


# ── Принять / выставить ────────────────────────────────────────────────────
@router.patch("/{company_id}/{esf_id}/accept")
def accept_esf(
    company_id: int,
    esf_id: int,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    esf = db.query(models.ESF).filter(
        models.ESF.id         == esf_id,
        models.ESF.company_id == company_id,
    ).first()
    if not esf:
        raise HTTPException(404, "ЭСФ не найден")
    esf.status      = "accepted" if esf.direction == "incoming" else "issued"
    esf.accepted_at = datetime.utcnow()
    db.commit()
    db.refresh(esf)
    return _esf_dict(esf)


# ── Отменить принятие ──────────────────────────────────────────────────────
@router.patch("/{company_id}/{esf_id}/unaccept")
def unaccept_esf(
    company_id: int,
    esf_id: int,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    esf = db.query(models.ESF).filter(
        models.ESF.id         == esf_id,
        models.ESF.company_id == company_id,
    ).first()
    if not esf:
        raise HTTPException(404, "ЭСФ не найден")
    esf.status      = "pending"
    esf.accepted_at = None
    db.commit()
    db.refresh(esf)
    return _esf_dict(esf)


# ── Привязать к банковской транзакции ─────────────────────────────────────
@router.patch("/{company_id}/{esf_id}/link-tx/{tx_id}")
def link_transaction(
    company_id: int,
    esf_id: int,
    tx_id: int,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    esf = db.query(models.ESF).filter(
        models.ESF.id         == esf_id,
        models.ESF.company_id == company_id,
    ).first()
    if not esf:
        raise HTTPException(404, "ЭСФ не найден")
    esf.bank_transaction_id = tx_id
    esf.linked_payment      = True
    db.commit()
    db.refresh(esf)
    return _esf_dict(esf)


# ── Отвязать от транзакции ────────────────────────────────────────────────
@router.patch("/{company_id}/{esf_id}/unlink-tx")
def unlink_transaction(
    company_id: int,
    esf_id: int,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    esf = db.query(models.ESF).filter(
        models.ESF.id         == esf_id,
        models.ESF.company_id == company_id,
    ).first()
    if not esf:
        raise HTTPException(404, "ЭСФ не найден")
    esf.bank_transaction_id = None
    esf.linked_payment      = False
    db.commit()
    db.refresh(esf)
    return _esf_dict(esf)


# ── Привязать к документу ─────────────────────────────────────────────────
@router.patch("/{company_id}/{esf_id}/link-doc/{doc_id}")
def link_document(
    company_id: int,
    esf_id: int,
    doc_id: int,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    esf = db.query(models.ESF).filter(
        models.ESF.id         == esf_id,
        models.ESF.company_id == company_id,
    ).first()
    if not esf:
        raise HTTPException(404, "ЭСФ не найден")
    esf.linked_document_id = doc_id
    db.commit()
    db.refresh(esf)
    return _esf_dict(esf)


# ── Экспорт XML для ИС ЭСФ ГНС КР ────────────────────────────────────────
@router.get("/{company_id}/export-xml")
def export_xml(
    company_id: int,
    direction:  str = "outgoing",
    date_from:  Optional[str] = None,
    date_to:    Optional[str] = None,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    """
    Генерирует XML-файл для импорта в ИС ЭСФ ГНС КР (esf.salyk.kg).
    Формат соответствует описанию XML-формата ГНС КР (sti.gov.kg, 2022).
    """
    import xml.etree.ElementTree as ET
    from xml.dom import minidom

    comp = db.query(models.Company).filter(models.Company.id == company_id).first()
    company_name = comp.name if comp else f"Компания #{company_id}"
    company_inn  = comp.inn  if comp else ""

    q = db.query(models.ESF).filter(
        models.ESF.company_id == company_id,
        models.ESF.direction  == direction,
    )
    if date_from:
        q = q.filter(models.ESF.esf_date >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        q = q.filter(models.ESF.esf_date <= dt_to)
    records = q.order_by(models.ESF.esf_date).all()

    # ── Строим XML ────────────────────────────────────────────────────────
    root = ET.Element("ЭСФСписок")
    root.set("версия", "1.0")
    root.set("датаФормирования", datetime.now().strftime("%Y-%m-%d"))
    root.set("организация", company_name)
    root.set("ИНН", company_inn or "")
    root.set("направление", "Исходящие" if direction == "outgoing" else "Входящие")

    VAT_RATE_LABEL = {"12": "12", "0": "0", "exempt": "БезНДС"}

    for r in records:
        esf_el = ET.SubElement(root, "ЭСФ")

        ET.SubElement(esf_el, "Номер").text = r.esf_number or ""
        ET.SubElement(esf_el, "Дата").text  = (
            r.esf_date.strftime("%Y-%m-%d") if r.esf_date else ""
        )
        if r.contract_number:
            ET.SubElement(esf_el, "НомерДоговора").text = r.contract_number

        # Продавец
        seller = ET.SubElement(esf_el, "Поставщик")
        ET.SubElement(seller, "ИНН").text          = r.supplier_inn  or ""
        ET.SubElement(seller, "Наименование").text = r.supplier_name or ""

        # Покупатель
        buyer = ET.SubElement(esf_el, "Покупатель")
        ET.SubElement(buyer, "ИНН").text          = r.buyer_inn  or ""
        ET.SubElement(buyer, "Наименование").text = r.buyer_name or ""

        # Суммы
        amount     = r.amount     or 0.0
        vat_amount = r.vat_amount or 0.0
        vat_rate   = VAT_RATE_LABEL.get(r.vat_rate or "12", r.vat_rate or "12")

        if vat_rate == "БезНДС":
            amount_no_vat = amount
            vat_amount    = 0.0
        elif vat_rate == "0":
            amount_no_vat = amount
            vat_amount    = 0.0
        else:
            # amount — это сумма С НДС
            amount_no_vat = round(amount - vat_amount, 2)

        ET.SubElement(esf_el, "СуммаСНДС").text    = f"{amount:.2f}"
        ET.SubElement(esf_el, "СтавкаНДС").text    = vat_rate
        ET.SubElement(esf_el, "СуммаНДС").text     = f"{vat_amount:.2f}"
        ET.SubElement(esf_el, "СуммаБезНДС").text  = f"{amount_no_vat:.2f}"

        # Статус
        STATUS_LABEL = {
            "pending":  "НеПринят",
            "accepted": "Принят",
            "issued":   "Выставлен",
        }
        ET.SubElement(esf_el, "Статус").text = STATUS_LABEL.get(r.status or "pending", r.status or "")

    # ── Красивое форматирование ───────────────────────────────────────────
    raw_xml = ET.tostring(root, encoding="unicode")
    pretty  = minidom.parseString(
        f'<?xml version="1.0" encoding="UTF-8"?>{raw_xml}'
    ).toprettyxml(indent="  ", encoding="UTF-8")

    direction_str = "iskhodyashchie" if direction == "outgoing" else "vkhodyashchie"
    filename = f"esf_{direction_str}_{company_id}"
    if date_from: filename += f"_ot{date_from}"
    if date_to:   filename += f"_do{date_to}"
    filename += ".xml"

    return StreamingResponse(
        io.BytesIO(pretty),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Отвязать от документа ─────────────────────────────────────────────────
@router.patch("/{company_id}/{esf_id}/unlink-doc")
def unlink_document(
    company_id: int,
    esf_id: int,
    db:   Session = Depends(get_db),
    company = Depends(require_company),
):
    esf = db.query(models.ESF).filter(
        models.ESF.id         == esf_id,
        models.ESF.company_id == company_id,
    ).first()
    if not esf:
        raise HTTPException(404, "ЭСФ не найден")
    esf.linked_document_id = None
    db.commit()
    db.refresh(esf)
    return _esf_dict(esf)
