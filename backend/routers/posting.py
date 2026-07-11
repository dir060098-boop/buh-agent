"""
AI-агент автоматической разноски + журнал хозяйственных операций КР (МСФО).
"""
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import and_
from database import get_db
from models import Document, JournalEntry, ChartOfAccount, PostingRule, Company
from routers.auth import get_current_user
from models import User
import anthropic, json, os
from datetime import date, datetime
from typing import Optional

router = APIRouter()
client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))


def get_chart_summary(db: Session) -> str:
    accounts = db.query(ChartOfAccount).filter(
        ChartOfAccount.level == 3, ChartOfAccount.is_active == True
    ).order_by(ChartOfAccount.code).all()
    if not accounts:
        return "План счетов не загружен."
    return "\n".join(f"{a.code} | {a.name} | {a.account_type}" for a in accounts)


def get_posting_rules_summary(db: Session) -> str:
    rules = db.query(PostingRule).filter(
        PostingRule.is_active == True
    ).order_by(PostingRule.priority.desc()).all()
    if not rules:
        return "Правила не загружены."
    lines = []
    for r in rules:
        kw = ", ".join(r.operation_keywords or [])
        lines.append(f"- {r.rule_name}: Дт {r.debit_account} / Кт {r.credit_account} | КС: {kw}")
    return "\n".join(lines)


class DuplicatePostingError(Exception):
    def __init__(self, entry: 'JournalEntry'):
        self.entry = entry

def post_document_with_ai(doc: Document, db: Session) -> JournalEntry:
    # ── ЗАЩИТА ОТ ПОВТОРНОЙ РАЗНОСКИ ──────────────────────────
    # Проверка 1: уже есть проводка для этого document_id
    existing_by_doc = db.query(JournalEntry).filter(
        JournalEntry.document_id == doc.id,
        JournalEntry.status.in_(["posted", "needs_review"])
    ).first()
    if existing_by_doc:
        raise DuplicatePostingError(existing_by_doc)

    # Проверка 2: аналогичный документ уже разнесён
    # (одинаковые: компания + сумма + валюта + дата + контрагент + тип)
    if doc.amount and doc.doc_date and doc.counterparty:
        existing_by_attrs = db.query(JournalEntry).join(
            Document, JournalEntry.document_id == Document.id
        ).filter(
            JournalEntry.company_id == doc.company_id,
            JournalEntry.status.in_(["posted", "needs_review"]),
            Document.amount == doc.amount,
            Document.currency == doc.currency,
            Document.counterparty == doc.counterparty,
            Document.doc_type == doc.doc_type,
            Document.id != doc.id
        ).first()
        if existing_by_attrs:
            raise DuplicatePostingError(existing_by_attrs)
    # ── КОНЕЦ ЗАЩИТЫ ──────────────────────────────────────────

    chart_summary = get_chart_summary(db)
    rules_summary = get_posting_rules_summary(db)

    prompt = f"""Ты — профессиональный бухгалтер КР, специалист МСФО.
Определи бухгалтерскую проводку для документа по плану счетов КР.

## ДОКУМЕНТ
Тип: {doc.doc_type}
Номер: {doc.doc_number or '—'}
Дата: {doc.doc_date or '—'}
Контрагент: {doc.counterparty or '—'}
ИНН контрагента: {doc.counterparty_inn or '—'}
Сумма: {doc.amount or 0} {doc.currency or 'KGS'}
НДС: {doc.vat_amount or 0}
Тип операции: {doc.operation_type or '—'}
Описание: {doc.ai_summary or doc.ai_raw_text or '—'}

## ПЛАН СЧЕТОВ КР (МСФО)
{chart_summary}

## ПРАВИЛА РАЗНОСКИ
{rules_summary}

Компания: торгово-импортная деятельность (КР).
Верни ТОЛЬКО JSON:
{{
  "debit_account": "XXXX",
  "credit_account": "XXXX",
  "debit_account_name": "название",
  "credit_account_name": "название",
  "amount": число,
  "currency": "KGS/RUB/USD/EUR/TRY",
  "description": "краткое содержание операции для журнала",
  "confidence": 0-100,
  "reasoning": "обоснование (1-2 предложения)",
  "needs_review": true/false
}}"""

    response = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=800,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = response.content[0].text.strip()

    try:
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        result = json.loads(raw)
    except Exception:
        result = {
            "debit_account": "7590", "credit_account": "3110",
            "debit_account_name": "Прочие операционные расходы",
            "credit_account_name": "Счета к оплате за товары и услуги",
            "amount": doc.amount or 0, "currency": doc.currency or "KGS",
            "description": "Не удалось разобрать ответ AI",
            "confidence": 0, "reasoning": "Ошибка парсинга", "needs_review": True
        }

    doc.debit_account = result.get("debit_account")
    doc.credit_account = result.get("credit_account")
    doc.ai_confidence = result.get("confidence", 0)
    doc.posting_status = "needs_review" if result.get("needs_review") else "posted"
    db.add(doc)

    # Конвертация в KGS только если валюта не KGS
    amount = result.get("amount", doc.amount or 0)
    currency = result.get("currency", doc.currency or "KGS")
    amount_kgs = float(amount) if currency == "KGS" else None  # TODO: курс НБКР

    entry = JournalEntry(
        company_id=doc.company_id,
        document_id=doc.id,
        scope=getattr(doc, "scope", None) or "official",
        entry_date=doc.doc_date.date() if doc.doc_date else date.today(),
        debit_account=result.get("debit_account"),
        credit_account=result.get("credit_account"),
        debit_account_name=result.get("debit_account_name"),
        credit_account_name=result.get("credit_account_name"),
        amount=amount,
        currency=currency,
        amount_kgs=amount_kgs,
        description=result.get("description"),
        ai_confidence=result.get("confidence", 0),
        ai_reasoning=result.get("reasoning"),
        status="needs_review" if result.get("needs_review") else "posted"
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry


# ── ENDPOINTS ────────────────────────────────────────────

@router.post("/auto/{document_id}")
def auto_post_document(document_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Документ не найден")
    company = db.query(Company).filter(Company.id == doc.company_id, Company.owner_id == current_user.id).first()
    if not company:
        raise HTTPException(status_code=403, detail="Нет доступа")
    try:
        entry = post_document_with_ai(doc, db)
    except DuplicatePostingError as e:
        raise HTTPException(status_code=409, detail={
            "error": "duplicate",
            "message": f"Документ уже разнесён (проводка #{e.entry.id})",
            "existing_entry_id": e.entry.id,
            "existing_status": e.entry.status,
            "debit": e.entry.debit_account,
            "credit": e.entry.credit_account,
            "amount": float(e.entry.amount),
            "currency": e.entry.currency
        })
    return {
        "success": True, "document_id": document_id, "entry_id": entry.id,
        "debit": f"{entry.debit_account} {entry.debit_account_name}",
        "credit": f"{entry.credit_account} {entry.credit_account_name}",
        "amount": float(entry.amount), "currency": entry.currency,
        "confidence": entry.ai_confidence, "status": entry.status,
        "description": entry.description, "reasoning": entry.ai_reasoning
    }


@router.post("/auto-all")
def auto_post_all(company_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    company = db.query(Company).filter(Company.id == company_id, Company.owner_id == current_user.id).first()
    if not company:
        raise HTTPException(status_code=403, detail="Нет доступа")
    docs = db.query(Document).filter(
        Document.company_id == company_id,
        Document.posting_status == "pending",
        Document.amount != None
    ).all()
    results, errors, skipped = [], [], []
    for doc in docs:
        try:
            entry = post_document_with_ai(doc, db)
            results.append({
                "document_id": doc.id, "doc_number": doc.doc_number,
                "debit": entry.debit_account, "credit": entry.credit_account,
                "amount": float(entry.amount), "currency": entry.currency,
                "confidence": entry.ai_confidence, "status": entry.status
            })
        except DuplicatePostingError as e:
            skipped.append({
                "document_id": doc.id,
                "doc_number": doc.doc_number,
                "reason": f"Уже разнесён (проводка #{e.entry.id}, статус: {e.entry.status})"
            })
        except Exception as e:
            errors.append({"document_id": doc.id, "error": str(e)})
    return {
        "processed": len(results),
        "skipped_duplicates": len(skipped),
        "errors": len(errors),
        "results": results,
        "skipped": skipped,
        "error_details": errors
    }


@router.get("/journal")
def get_journal(
    company_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    counterparty: Optional[str] = None,
    debit_account: Optional[str] = None,
    scope: Optional[str] = None,
    include_archived: bool = False,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Журнал хозяйственных операций с полными реквизитами."""
    company = db.query(Company).filter(Company.id == company_id, Company.owner_id == current_user.id).first()
    if not company:
        raise HTTPException(status_code=403, detail="Нет доступа")

    q = db.query(JournalEntry, Document).outerjoin(
        Document, JournalEntry.document_id == Document.id
    ).filter(JournalEntry.company_id == company_id)

    # По умолчанию скрываем архивные (закрытые периоды)
    if not include_archived:
        q = q.filter(JournalEntry.is_archived == False)

    if date_from:
        q = q.filter(JournalEntry.entry_date >= date_from)
    if date_to:
        q = q.filter(JournalEntry.entry_date <= date_to)
    if status:
        q = q.filter(JournalEntry.status == status)
    if debit_account:
        q = q.filter(JournalEntry.debit_account == debit_account)
    if scope in ("official", "internal"):
        q = q.filter(JournalEntry.scope == scope)

    # Фильтр по контрагенту (через JOIN с Document)
    if counterparty:
        q = q.filter(Document.counterparty.ilike(f"%{counterparty}%"))

    total = q.count()
    rows  = q.order_by(JournalEntry.entry_date.desc(), JournalEntry.id.desc())\
             .offset(offset).limit(limit).all()

    result = []
    for i, (e, doc) in enumerate(rows):
        result.append({
            "row_num": len(rows) - i,          # № п/п (обратный порядок)
            "id": e.id,
            "entry_date": str(e.entry_date),
            # Реквизиты документа-основания
            "doc_number": doc.doc_number if doc else None,
            "doc_type": doc.doc_type if doc else None,
            "doc_date": str(doc.doc_date)[:10] if doc and doc.doc_date else None,
            "counterparty": doc.counterparty if doc else None,
            "counterparty_inn": doc.counterparty_inn if doc else None,
            # Бухгалтерские данные
            "description": e.description,
            "debit_account": e.debit_account,
            "debit_account_name": e.debit_account_name,
            "credit_account": e.credit_account,
            "credit_account_name": e.credit_account_name,
            "amount": float(e.amount),
            "currency": e.currency,
            "amount_kgs": float(e.amount_kgs) if e.amount_kgs else (float(e.amount) if e.currency == "KGS" else None),
            "exchange_rate": float(e.exchange_rate) if e.exchange_rate else None,
            # AI данные
            "ai_confidence": e.ai_confidence,
            "ai_reasoning": e.ai_reasoning,
            "scope": e.scope or "official",
            "status": e.status,
            "is_archived": e.is_archived or False,
            "archived_at": str(e.archived_at) if e.archived_at else None,
            "document_id": e.document_id,
            "reviewed_by": e.reviewed_by,
            "reviewed_at": str(e.reviewed_at) if e.reviewed_at else None,
            "created_at": str(e.created_at)
        })
    return {"items": result, "total": total, "has_more": offset + limit < total}


# ── Предпросмотр закрытия периода ─────────────────────────────────────────
@router.get("/period-preview")
def period_preview(
    company_id: int,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Сколько posted-проводок будет заархивировано за выбранный месяц."""
    company = db.query(Company).filter(
        Company.id == company_id, Company.owner_id == current_user.id
    ).first()
    if not company:
        raise HTTPException(403, "Нет доступа")

    from_date = date(year, month, 1)
    to_date   = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    count = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.entry_date >= from_date,
        JournalEntry.entry_date <  to_date,
        JournalEntry.status     == "posted",
        JournalEntry.is_archived == False,
    ).count()

    MONTHS_RU = ["","Январь","Февраль","Март","Апрель","Май","Июнь",
                 "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    return {
        "year": year, "month": month,
        "period_label": f"{MONTHS_RU[month]} {year}",
        "count": count,
    }


# ── Закрыть период (архивировать) ─────────────────────────────────────────
class ClosePeriodRequest(BaseModel):
    year: int
    month: int   # 1–12

@router.post("/close-period")
def close_period(
    company_id: int,
    data: ClosePeriodRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Архивирует все posted-проводки за выбранный месяц."""
    company = db.query(Company).filter(
        Company.id == company_id, Company.owner_id == current_user.id
    ).first()
    if not company:
        raise HTTPException(403, "Нет доступа")

    from_date = date(data.year, data.month, 1)
    to_date   = date(data.year + 1, 1, 1) if data.month == 12 else date(data.year, data.month + 1, 1)

    entries = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.entry_date >= from_date,
        JournalEntry.entry_date <  to_date,
        JournalEntry.status     == "posted",
        JournalEntry.is_archived == False,
    ).all()

    now = datetime.utcnow()
    for e in entries:
        e.is_archived = True
        e.archived_at = now
    db.commit()

    MONTHS_RU = ["","Январь","Февраль","Март","Апрель","Май","Июнь",
                 "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    return {
        "archived": len(entries),
        "period_label": f"{MONTHS_RU[data.month]} {data.year}",
    }


# ── Переоткрыть период ────────────────────────────────────────────────────
@router.post("/reopen-period")
def reopen_period(
    company_id: int,
    data: ClosePeriodRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Снимает архивацию с проводок выбранного месяца (отмена закрытия)."""
    company = db.query(Company).filter(
        Company.id == company_id, Company.owner_id == current_user.id
    ).first()
    if not company:
        raise HTTPException(403, "Нет доступа")

    from_date = date(data.year, data.month, 1)
    to_date   = date(data.year + 1, 1, 1) if data.month == 12 else date(data.year, data.month + 1, 1)

    entries = db.query(JournalEntry).filter(
        JournalEntry.company_id  == company_id,
        JournalEntry.entry_date  >= from_date,
        JournalEntry.entry_date  <  to_date,
        JournalEntry.is_archived == True,
    ).all()

    for e in entries:
        e.is_archived = False
        e.archived_at = None
    db.commit()

    MONTHS_RU = ["","Январь","Февраль","Март","Апрель","Май","Июнь",
                 "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    return {
        "reopened": len(entries),
        "period_label": f"{MONTHS_RU[data.month]} {data.year}",
    }


# ── Серверная статистика журнала (вся выборка, не страница) ──────────────
@router.get("/journal-stats")
def journal_stats(
    company_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    include_archived: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Итоги по всему журналу с учётом фильтров (не зависят от пагинации)."""
    company = db.query(Company).filter(
        Company.id == company_id, Company.owner_id == current_user.id
    ).first()
    if not company:
        raise HTTPException(403, "Нет доступа")

    from sqlalchemy import func as sa_func
    q = db.query(JournalEntry).filter(JournalEntry.company_id == company_id)
    if not include_archived:
        q = q.filter(JournalEntry.is_archived == False)
    if date_from:
        q = q.filter(JournalEntry.entry_date >= date_from)
    if date_to:
        q = q.filter(JournalEntry.entry_date <= date_to)

    total        = q.count()
    posted       = q.filter(JournalEntry.status == "posted").count()
    # Пересоздаём базовый запрос т.к. q уже отфильтрован по posted
    q2 = db.query(JournalEntry).filter(JournalEntry.company_id == company_id)
    if not include_archived:
        q2 = q2.filter(JournalEntry.is_archived == False)
    if date_from:
        q2 = q2.filter(JournalEntry.entry_date >= date_from)
    if date_to:
        q2 = q2.filter(JournalEntry.entry_date <= date_to)
    needs_review = q2.filter(JournalEntry.status == "needs_review").count()

    sum_q = db.query(sa_func.coalesce(sa_func.sum(JournalEntry.amount), 0)).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.status     == "posted",
        JournalEntry.currency   == "KGS",
    )
    if not include_archived:
        sum_q = sum_q.filter(JournalEntry.is_archived == False)
    if date_from:
        sum_q = sum_q.filter(JournalEntry.entry_date >= date_from)
    if date_to:
        sum_q = sum_q.filter(JournalEntry.entry_date <= date_to)
    total_kgs = float(sum_q.scalar() or 0)

    return {
        "total":        total,
        "posted":       posted,
        "needs_review": needs_review,
        "total_kgs":    round(total_kgs, 2),
    }


# ── ОСВ: Оборотно-сальдовая ведомость ────────────────────────────────────
@router.get("/trial-balance")
def trial_balance(
    company_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    scope: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Оборотно-сальдовая ведомость за период.

    Для каждого счёта: сальдо начальное (Дт/Кт), обороты за период (Дт/Кт),
    сальдо конечное (Дт/Кт). Учитываются ВСЕ posted-проводки включая архив.
    Сумма в KGS: amount_kgs если есть, иначе amount (для KGS-проводок).
    scope: official | internal | None (всё) — фильтр контура учёта.
    """
    company = db.query(Company).filter(
        Company.id == company_id, Company.owner_id == current_user.id
    ).first()
    if not company:
        raise HTTPException(403, "Нет доступа")

    from sqlalchemy import func as sa_func, case

    # Сумма в KGS: COALESCE(amount_kgs, amount)
    amt = sa_func.coalesce(JournalEntry.amount_kgs, JournalEntry.amount)

    def _turnovers(account_col, dt_from=None, dt_to=None):
        """Обороты по счёту (дебетовые или кредитовые) за интервал."""
        q = db.query(
            account_col.label("account"),
            sa_func.sum(amt).label("total"),
        ).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.status     == "posted",
        )
        if scope in ("official", "internal"):
            q = q.filter(JournalEntry.scope == scope)
        if dt_from:
            q = q.filter(JournalEntry.entry_date >= dt_from)
        if dt_to:
            q = q.filter(JournalEntry.entry_date <= dt_to)
        return {r.account: float(r.total or 0) for r in q.group_by(account_col).all()}

    # Обороты ДО периода (для сальдо начального)
    open_debit, open_credit = {}, {}
    if date_from:
        from datetime import date as _date, timedelta as _td
        day_before = (_date.fromisoformat(date_from) - _td(days=1)).isoformat()
        open_debit  = _turnovers(JournalEntry.debit_account,  dt_to=day_before)
        open_credit = _turnovers(JournalEntry.credit_account, dt_to=day_before)

    # Обороты ЗА период
    per_debit  = _turnovers(JournalEntry.debit_account,  dt_from=date_from, dt_to=date_to)
    per_credit = _turnovers(JournalEntry.credit_account, dt_from=date_from, dt_to=date_to)

    # Все счета встречающиеся в данных
    accounts = sorted(
        set(open_debit) | set(open_credit) | set(per_debit) | set(per_credit)
    )
    if not accounts:
        return {"rows": [], "totals": {}, "period": {"from": date_from, "to": date_to}}

    # Названия счетов из плана
    chart = {
        a.code: a.name
        for a in db.query(ChartOfAccount).filter(ChartOfAccount.code.in_(accounts)).all()
    }

    rows = []
    t_ob_d = t_ob_k = t_per_d = t_per_k = t_cb_d = t_cb_k = 0.0

    for code in accounts:
        # Сальдо начальное: дебетовые обороты − кредитовые (до периода)
        ob_net = open_debit.get(code, 0) - open_credit.get(code, 0)
        ob_d = round(ob_net, 2)  if ob_net > 0 else 0.0
        ob_k = round(-ob_net, 2) if ob_net < 0 else 0.0

        pd_ = round(per_debit.get(code, 0), 2)
        pk_ = round(per_credit.get(code, 0), 2)

        # Сальдо конечное
        cb_net = ob_net + pd_ - pk_
        cb_d = round(cb_net, 2)  if cb_net > 0 else 0.0
        cb_k = round(-cb_net, 2) if cb_net < 0 else 0.0

        # Пропускаем полностью нулевые строки
        if not any([ob_d, ob_k, pd_, pk_, cb_d, cb_k]):
            continue

        rows.append({
            "account":       code,
            "account_name":  chart.get(code, ""),
            "opening_debit":  ob_d,
            "opening_credit": ob_k,
            "period_debit":   pd_,
            "period_credit":  pk_,
            "closing_debit":  cb_d,
            "closing_credit": cb_k,
        })
        t_ob_d += ob_d; t_ob_k += ob_k
        t_per_d += pd_; t_per_k += pk_
        t_cb_d += cb_d; t_cb_k += cb_k

    return {
        "rows": rows,
        "totals": {
            "opening_debit":  round(t_ob_d, 2),
            "opening_credit": round(t_ob_k, 2),
            "period_debit":   round(t_per_d, 2),
            "period_credit":  round(t_per_k, 2),
            "closing_debit":  round(t_cb_d, 2),
            "closing_credit": round(t_cb_k, 2),
        },
        "period": {"from": date_from, "to": date_to},
        "balanced": abs(t_per_d - t_per_k) < 0.01,
    }


# ── ОСВ: экспорт в Excel ──────────────────────────────────────────────────
@router.get("/trial-balance/export")
def trial_balance_export(
    company_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    scope: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """ОСВ в Excel — тот же расчёт, что и на экране."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    from datetime import datetime as _dt
    import io as _io

    data = trial_balance(company_id, date_from, date_to, scope, db, current_user)
    company = db.query(Company).filter(Company.id == company_id).first()
    company_name = company.name if company else f"Компания #{company_id}"

    SCOPE_RU = {"official": "официальный контур", "internal": "внутренний контур"}
    period_str = f"{date_from or '…'} — {date_to or '…'}"
    scope_str = f" · {SCOPE_RU[scope]}" if scope in SCOPE_RU else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "ОСВ"

    hdr_fill  = PatternFill("solid", fgColor="1A56DB")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_aln   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sub_fill  = PatternFill("solid", fgColor="EEF2FF")
    cell_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", bold=True, size=10)
    total_fill = PatternFill("solid", fgColor="DBEAFE")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{company_name} — Оборотно-сальдовая ведомость · {period_str}{scope_str}"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    balanced_str = "" if data["balanced"] else " · ⚠ ОБОРОТЫ ДТ ≠ КТ — проверьте проводки"
    ws["A2"] = f"Сформировано: {_dt.now().strftime('%d.%m.%Y %H:%M')} · в сомах (KGS){balanced_str}"
    ws["A2"].font = Font(name="Arial", size=9,
                         color="CC0000" if not data["balanced"] else "888888")

    # Двухуровневая шапка
    ws.merge_cells("A3:B3"); ws.merge_cells("C3:D3")
    ws.merge_cells("E3:F3"); ws.merge_cells("G3:H3")
    for rng, title in [("A3", "Счёт"), ("C3", "Сальдо начальное"),
                       ("E3", "Обороты за период"), ("G3", "Сальдо конечное")]:
        c = ws[rng]
        c.value = title
        c.font, c.fill, c.alignment = hdr_font, hdr_fill, hdr_aln
    sub_headers = ["Код", "Название", "Дт", "Кт", "Дт", "Кт", "Дт", "Кт"]
    for col, h in enumerate(sub_headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_aln, border
    for row in (3, 4):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border

    for idx, r in enumerate(data["rows"], 1):
        row_n = idx + 4
        vals = [r["account"], r["account_name"],
                r["opening_debit"], r["opening_credit"],
                r["period_debit"], r["period_credit"],
                r["closing_debit"], r["closing_credit"]]
        fill = sub_fill if idx % 2 == 0 else None
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_n, column=col,
                           value=(val if not (col > 2 and not val) else None))
            cell.font, cell.border = cell_font, border
            if fill:
                cell.fill = fill
            if col > 2:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    t = data["totals"]
    total_row = len(data["rows"]) + 5
    ws.merge_cells(f"A{total_row}:B{total_row}")
    ws.cell(total_row, 1, "ИТОГО").alignment = Alignment(horizontal="right")
    for col, val in enumerate([None, None, t["opening_debit"], t["opening_credit"],
                               t["period_debit"], t["period_credit"],
                               t["closing_debit"], t["closing_credit"]], 1):
        cell = ws.cell(total_row, col)
        if val is not None:
            cell.value = val
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        cell.font, cell.fill, cell.border = total_font, total_fill, border

    for i, w in enumerate([9, 38, 14, 14, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"osv_{company_id}"
    if date_from: filename += f"_{date_from}"
    if date_to:   filename += f"_{date_to}"
    if scope:     filename += f"_{scope}"
    filename += ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Список закрытых периодов компании ────────────────────────────────────
@router.get("/closed-periods")
def get_closed_periods(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Возвращает список месяцев с архивными проводками и их количество."""
    company = db.query(Company).filter(
        Company.id == company_id, Company.owner_id == current_user.id
    ).first()
    if not company:
        raise HTTPException(403, "Нет доступа")

    from sqlalchemy import func, extract
    rows = db.query(
        extract("year",  JournalEntry.entry_date).label("year"),
        extract("month", JournalEntry.entry_date).label("month"),
        func.count(JournalEntry.id).label("count"),
    ).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.is_archived == True,
    ).group_by("year", "month").order_by("year", "month").all()

    MONTHS_RU = ["","Январь","Февраль","Март","Апрель","Май","Июнь",
                 "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    return [
        {
            "year": int(r.year), "month": int(r.month),
            "period_label": f"{MONTHS_RU[int(r.month)]} {int(r.year)}",
            "count": r.count,
        }
        for r in rows
    ]


@router.get("/chart-of-accounts")
def get_chart_of_accounts(level: Optional[int] = None, section: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(ChartOfAccount).filter(ChartOfAccount.is_active == True)
    if level:
        q = q.filter(ChartOfAccount.level == level)
    if section:
        q = q.filter(ChartOfAccount.section == section)
    accounts = q.order_by(ChartOfAccount.code).all()
    return [{"code": a.code, "name": a.name, "section": a.section, "account_type": a.account_type, "level": a.level, "parent_code": a.parent_code} for a in accounts]


@router.post("/seed-chart")
def seed_chart_of_accounts(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Добавляет только отсутствующие счета (не трогает существующие)."""
    from seed_chart import CHART_OF_ACCOUNTS, POSTING_RULES
    loaded_accounts = 0
    for item in CHART_OF_ACCOUNTS:
        existing = db.query(ChartOfAccount).filter(ChartOfAccount.code == item["code"]).first()
        if existing:
            # Обновляем название если изменилось
            existing.name = item["name"]
        else:
            db.add(ChartOfAccount(**item))
            loaded_accounts += 1
    loaded_rules = 0
    for item in POSTING_RULES:
        existing = db.query(PostingRule).filter(PostingRule.rule_name == item["rule_name"]).first()
        if existing:
            existing.debit_account  = item["debit_account"]
            existing.credit_account = item["credit_account"]
            existing.description    = item["description"]
            existing.priority       = item["priority"]
            existing.operation_keywords = item["operation_keywords"]
        else:
            db.add(PostingRule(**item))
            loaded_rules += 1
    db.commit()
    total_accounts = db.query(ChartOfAccount).count()
    total_rules    = db.query(PostingRule).count()
    return {
        "success": True,
        "accounts_added": loaded_accounts,
        "rules_added": loaded_rules,
        "total_accounts": total_accounts,
        "total_rules": total_rules
    }


@router.post("/reseed-chart")
def reseed_chart(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """ПОЛНАЯ перезагрузка плана счетов — удаляет все старые и загружает новые.
    Вызывать при смене официального плана счетов."""
    from seed_chart import CHART_OF_ACCOUNTS, POSTING_RULES
    # Удаляем только правила разноски и счета (не трогаем проводки)
    db.query(PostingRule).delete()
    db.query(ChartOfAccount).delete()
    db.commit()
    for item in CHART_OF_ACCOUNTS:
        db.add(ChartOfAccount(**item))
    for item in POSTING_RULES:
        db.add(PostingRule(**item))
    db.commit()
    return {
        "success": True,
        "accounts_loaded": len(CHART_OF_ACCOUNTS),
        "rules_loaded": len(POSTING_RULES),
        "message": f"План счетов КР обновлён: {len(CHART_OF_ACCOUNTS)} счетов, {len(POSTING_RULES)} правил разноски"
    }


# ── ФЛОУ "НА ПРОВЕРКЕ" ──────────────────────────────────

class ReviewAction(BaseModel):
    action: str  # confirm | reject | correct
    debit_account: Optional[str] = None
    credit_account: Optional[str] = None
    description: Optional[str] = None
    comment: Optional[str] = None  # причина отклонения

@router.patch("/journal/{entry_id}/review")
def review_entry(
    entry_id: int,
    data: ReviewAction,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Флоу проверки проводки бухгалтером:
    - confirm  → статус posted, фиксируем кто подтвердил
    - reject   → статус rejected, фиксируем причину
    - correct  → меняем Дт/Кт, статус posted
    """
    entry = db.query(JournalEntry).filter(JournalEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Проводка не найдена")

    # Проверяем доступ через компанию
    company = db.query(Company).filter(
        Company.id == entry.company_id,
        Company.owner_id == current_user.id
    ).first()
    if not company:
        raise HTTPException(status_code=403, detail="Нет доступа")

    now = datetime.utcnow()
    reviewer = current_user.email

    if data.action == "confirm":
        entry.status = "posted"
        entry.reviewed_by = reviewer
        entry.reviewed_at = now

    elif data.action == "reject":
        entry.status = "rejected"
        entry.reviewed_by = reviewer
        entry.reviewed_at = now
        if data.comment:
            entry.ai_reasoning = f"[Отклонено: {data.comment}] " + (entry.ai_reasoning or "")
        # Обновляем статус документа
        if entry.document_id:
            doc = db.query(Document).filter(Document.id == entry.document_id).first()
            if doc:
                doc.posting_status = "rejected"
                db.add(doc)

    elif data.action == "correct":
        # Проверяем что счета существуют
        if data.debit_account:
            acc = db.query(ChartOfAccount).filter(ChartOfAccount.code == data.debit_account).first()
            if not acc:
                raise HTTPException(status_code=400, detail=f"Счёт {data.debit_account} не найден в плане счетов КР")
            entry.debit_account = data.debit_account
            entry.debit_account_name = acc.name
        if data.credit_account:
            acc = db.query(ChartOfAccount).filter(ChartOfAccount.code == data.credit_account).first()
            if not acc:
                raise HTTPException(status_code=400, detail=f"Счёт {data.credit_account} не найден в плане счетов КР")
            entry.credit_account = data.credit_account
            entry.credit_account_name = acc.name
        if data.description:
            entry.description = data.description
        entry.status = "posted"
        entry.reviewed_by = reviewer
        entry.reviewed_at = now
        # Обновляем документ
        if entry.document_id:
            doc = db.query(Document).filter(Document.id == entry.document_id).first()
            if doc:
                doc.debit_account = entry.debit_account
                doc.credit_account = entry.credit_account
                doc.posting_status = "posted"
                db.add(doc)
    else:
        raise HTTPException(status_code=400, detail="action должен быть: confirm | reject | correct")

    db.add(entry)
    db.commit()
    db.refresh(entry)

    return {
        "id": entry.id,
        "status": entry.status,
        "debit_account": entry.debit_account,
        "debit_account_name": entry.debit_account_name,
        "credit_account": entry.credit_account,
        "credit_account_name": entry.credit_account_name,
        "reviewed_by": entry.reviewed_by,
        "reviewed_at": str(entry.reviewed_at)
    }


@router.delete("/journal/{entry_id}")
def delete_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Удалить проводку из журнала (для исправления дублей)."""
    entry = db.query(JournalEntry).filter(JournalEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Проводка не найдена")

    company = db.query(Company).filter(
        Company.id == entry.company_id,
        Company.owner_id == current_user.id
    ).first()
    if not company:
        raise HTTPException(status_code=403, detail="Нет доступа")

    # Сбрасываем статус документа обратно в pending
    if entry.document_id:
        doc = db.query(Document).filter(Document.id == entry.document_id).first()
        if doc:
            # Проверяем не осталось ли других проводок для этого документа
            other_entries = db.query(JournalEntry).filter(
                JournalEntry.document_id == doc.id,
                JournalEntry.id != entry_id
            ).count()
            if other_entries == 0:
                doc.posting_status = "pending"
                doc.debit_account = None
                doc.credit_account = None
                db.add(doc)

    db.delete(entry)
    db.commit()
    return {"ok": True, "deleted_entry_id": entry_id}


@router.post("/journal/bulk-delete")
def bulk_delete_entries(
    entry_ids: list[int],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Удалить несколько проводок сразу (для очистки дублей)."""
    deleted = []
    errors = []

    for entry_id in entry_ids:
        entry = db.query(JournalEntry).filter(JournalEntry.id == entry_id).first()
        if not entry:
            errors.append({"id": entry_id, "error": "не найдена"})
            continue

        company = db.query(Company).filter(
            Company.id == entry.company_id,
            Company.owner_id == current_user.id
        ).first()
        if not company:
            errors.append({"id": entry_id, "error": "нет доступа"})
            continue

        if entry.document_id:
            doc = db.query(Document).filter(Document.id == entry.document_id).first()
            if doc:
                other = db.query(JournalEntry).filter(
                    JournalEntry.document_id == doc.id,
                    JournalEntry.id != entry_id
                ).count()
                if other == 0:
                    doc.posting_status = "pending"
                    doc.debit_account = None
                    doc.credit_account = None
                    db.add(doc)

        db.delete(entry)
        deleted.append(entry_id)

    db.commit()
    return {"deleted": len(deleted), "errors": len(errors), "deleted_ids": deleted, "error_details": errors}
