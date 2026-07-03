from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_
from database import get_db
from routers.auth import get_current_user, require_company
from typing import Optional
import models

router = APIRouter()

DOC_TYPE_LABELS = {
    "invoice": "Счёт на оплату",
    "act": "Акт",
    "esf": "ЭСФ",
    "ttn": "Накладная (ТТН)",
    "contract": "Договор",
    "receipt": "Квитанция",
    "payment_order": "Платёжное поручение",
    "bank_statement": "Выписка банка",
    "payroll": "Зарплатная ведомость",
    "other": "Прочее",
}

def doc_to_dict(doc):
    return {
        "id": doc.id,
        "doc_type": doc.doc_type.value if hasattr(doc.doc_type, "value") else str(doc.doc_type),
        "doc_type_label": DOC_TYPE_LABELS.get(
            doc.doc_type.value if hasattr(doc.doc_type, "value") else str(doc.doc_type), "Прочее"
        ),
        "doc_number": doc.doc_number,
        "doc_date": str(doc.doc_date)[:10] if doc.doc_date else None,
        "counterparty": doc.counterparty,
        "counterparty_inn": doc.counterparty_inn,
        "amount": doc.amount,
        "currency": doc.currency or "KGS",
        "vat_amount": doc.vat_amount or 0,
        "posting_status": doc.posting_status or "pending",
        "operation_type": doc.operation_type,
        "ai_confidence": doc.ai_confidence,
        "ai_summary": doc.ai_summary,
        "file_path": doc.file_path,
        "created_at": doc.created_at.isoformat() if doc.created_at else None,
    }


@router.get("/{company_id}")
def list_documents(
    company_id: int,
    search: Optional[str] = Query(None),
    doc_type: Optional[str] = Query(None),
    posting_status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    company = Depends(require_company),
):
    q = db.query(models.Document).filter(models.Document.company_id == company_id)

    if search:
        q = q.filter(
            or_(
                models.Document.counterparty.ilike(f"%{search}%"),
                models.Document.doc_number.ilike(f"%{search}%"),
                models.Document.operation_type.ilike(f"%{search}%"),
            )
        )
    if doc_type:
        q = q.filter(models.Document.doc_type == doc_type)
    if posting_status:
        q = q.filter(models.Document.posting_status == posting_status)
    if date_from:
        q = q.filter(models.Document.doc_date >= date_from)
    if date_to:
        q = q.filter(models.Document.doc_date <= date_to)

    total = q.count()
    docs  = q.order_by(models.Document.created_at.desc()).offset(offset).limit(limit).all()
    return {"items": [doc_to_dict(d) for d in docs], "total": total, "has_more": offset + limit < total}


@router.get("/{company_id}/export-1c")
def export_1c(
    company_id: int,
    search: Optional[str] = Query(None),
    doc_type: Optional[str] = Query(None),
    posting_status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    company = Depends(require_company),
):
    """
    Выгружает первичные документы в Excel под универсальную загрузку
    в «1С:Бухгалтерия 8.3» (обработка «Загрузка данных из табличного
    документа»): один документ = одна строка-услуга/товар (без разбивки
    по номенклатуре — БухАгент не ведёт складской учёт по позициям).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    from datetime import datetime as _dt
    import io as _io

    comp = db.query(models.Company).filter(models.Company.id == company_id).first()
    company_name = comp.name if comp else f"Компания #{company_id}"

    q = db.query(models.Document).filter(models.Document.company_id == company_id)
    if search:
        q = q.filter(or_(
            models.Document.counterparty.ilike(f"%{search}%"),
            models.Document.doc_number.ilike(f"%{search}%"),
            models.Document.operation_type.ilike(f"%{search}%"),
        ))
    if doc_type:
        q = q.filter(models.Document.doc_type == doc_type)
    if posting_status:
        q = q.filter(models.Document.posting_status == posting_status)
    if date_from:
        q = q.filter(models.Document.doc_date >= date_from)
    if date_to:
        q = q.filter(models.Document.doc_date <= date_to)

    docs = q.order_by(models.Document.doc_date).all()
    if not docs:
        raise HTTPException(404, "Нет документов для выгрузки за указанный период")

    wb = Workbook()
    ws = wb.active
    ws.title = "Документы"

    accent    = "1A56DB"
    hdr_fill  = PatternFill("solid", fgColor=accent)
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_aln   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sub_fill  = PatternFill("solid", fgColor="EEF2FF")
    cell_font = Font(name="Arial", size=10)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:P1")
    ws["A1"] = f"{company_name} — Документы для загрузки в 1С"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:P2")
    ws["A2"] = f"Сформировано: {_dt.now().strftime('%d.%m.%Y %H:%M')} · документов: {len(docs)}"
    ws["A2"].font = Font(name="Arial", size=9, color="888888")
    ws.row_dimensions[2].height = 14

    headers = [
        "№", "Дата документа", "Номер документа", "Вид операции",
        "Контрагент", "ИНН контрагента", "Номенклатура (содержание)",
        "Кол-во", "Цена (без НДС)", "Сумма (с НДС)",
        "Ставка НДС", "Сумма НДС", "Валюта",
        "Счёт Дт", "Счёт Кт", "Статус разноски",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_aln, border
    ws.row_dimensions[3].height = 32

    STATUS_RU = {"pending": "Не разнесён", "posted": "Разнесён", "needs_review": "На проверке"}

    for idx, d in enumerate(docs, 1):
        row_n = idx + 3
        amount = d.amount or 0.0
        vat    = d.vat_amount or 0.0
        price_no_vat = round(amount - vat, 2)
        if vat > 0 and (amount - vat) > 0:
            vat_rate_label = f"{round(vat / (amount - vat) * 100)}%"
        else:
            vat_rate_label = "Без НДС"
        doc_type_val = d.doc_type.value if hasattr(d.doc_type, "value") else str(d.doc_type)
        nomenclature = d.operation_type or d.ai_summary or DOC_TYPE_LABELS.get(doc_type_val, "Прочее")

        row_data = [
            idx,
            d.doc_date.strftime("%d.%m.%Y") if d.doc_date else "—",
            d.doc_number or "—",
            DOC_TYPE_LABELS.get(doc_type_val, "Прочее"),
            d.counterparty or "—",
            d.counterparty_inn or "—",
            nomenclature[:200] if nomenclature else "—",
            1,
            price_no_vat,
            round(amount, 2),
            vat_rate_label,
            round(vat, 2),
            d.currency or "KGS",
            d.debit_account or "—",
            d.credit_account or "—",
            STATUS_RU.get(d.posting_status or "pending", d.posting_status or "—"),
        ]
        fill = sub_fill if idx % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            cell.font, cell.border = cell_font, border
            if fill:
                cell.fill = fill
            if col in (9, 10, 12):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")

    widths = [5, 14, 16, 16, 30, 16, 32, 7, 13, 13, 10, 12, 8, 9, 9, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"documents_1c_{company_id}"
    if date_from: filename += f"_от{date_from}"
    if date_to:   filename += f"_до{date_to}"
    filename += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/doc/{document_id}")
def get_document(document_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Документ не найден")
    company = db.query(models.Company).filter(
        models.Company.id == doc.company_id,
        models.Company.owner_id == user.id,
    ).first()
    if not company:
        raise HTTPException(status_code=403, detail="Нет доступа")
    return doc_to_dict(doc)


@router.delete("/{doc_id}")
def delete_document(doc_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    doc = (
        db.query(models.Document)
        .join(models.Company, models.Document.company_id == models.Company.id)
        .filter(models.Document.id == doc_id, models.Company.owner_id == user.id)
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Документ не найден")

    # Каскад: очищаем ссылки на документ в ЭСФ и банковских транзакциях
    db.query(models.ESF).filter(
        models.ESF.linked_document_id == doc_id
    ).update({"linked_document_id": None})
    db.query(models.BankTransaction).filter(
        models.BankTransaction.linked_document_id == doc_id
    ).update({"linked_document_id": None, "status": "unmatched"})

    db.delete(doc)
    db.commit()
    return {"ok": True}
